import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Función para extraer líneas de productos de texto directo
def extraer_lineas_productos(texto):
    lineas = texto.split('\n')
    productos = []
    for linea in lineas:
        match = re.search(r'(\d+[xX]\d+[xX]\d+)', linea)
        if match:
            numeros = re.findall(r'\d+(?:[.,]\d+)?', linea)
            if len(numeros) >= 2:
                descripcion = linea.strip()
                for i in range(len(numeros) - 1):
                    cantidad_cand = float(numeros[i].replace(',', '.'))
                    precio_cand = float(numeros[i+1].replace(',', '.'))
                    if 0 < precio_cand < 1000:
                        cantidad = cantidad_cand
                        precio = precio_cand
                        break
                productos.append({
                    "Producto": descripcion,
                    "Cantidad": cantidad,
                    "Precio Unitario": precio,
                    "Dimensiones": tuple(sorted(map(lambda x: float(x.replace(',', '.')), match.group(1).lower().split('x'))))
                })
    return productos

# Función para agrupar productos por dimensión
def agrupar_por_dimension(df):
    return df.groupby("Dimensiones").agg({
        "Producto": lambda x: " + ".join(x),
        "Cantidad": "sum",
        "Precio Unitario": "mean"
    }).reset_index()

# Función para comparar dos dataframes agrupados
def comparar_factura_pedido(factura_df, pedido_df):
    resultados = []
    for _, f_row in factura_df.iterrows():
        match = pedido_df[pedido_df["Dimensiones"] == f_row["Dimensiones"]]
        if not match.empty:
            p_row = match.iloc[0]
            cantidad_igual = f_row["Cantidad"] == p_row["Cantidad"]
            precio_igual = abs(f_row["Precio Unitario"] - p_row["Precio Unitario"]) < 0.001
            if cantidad_igual and precio_igual:
                estado = "Correcto"
            elif cantidad_igual:
                estado = "Precio diferente"
            elif precio_igual:
                estado = "Cantidad diferente"
            else:
                estado = "Ambos diferentes"
            resultados.append({
                "Producto Factura": f_row["Producto"],
                "Cantidad Factura": f_row["Cantidad"],
                "Precio Factura": f_row["Precio Unitario"],
                "Producto Pedido": p_row["Producto"],
                "Cantidad Pedido": p_row["Cantidad"],
                "Precio Pedido": p_row["Precio Unitario"],
                "Dimensiones": "x".join(map(str, f_row["Dimensiones"])),
                "Estado": estado
            })
        else:
            resultados.append({
                "Producto Factura": f_row["Producto"],
                "Cantidad Factura": f_row["Cantidad"],
                "Precio Factura": f_row["Precio Unitario"],
                "Producto Pedido": "NO ENCONTRADO",
                "Cantidad Pedido": "",
                "Precio Pedido": "",
                "Dimensiones": "x".join(map(str, f_row["Dimensiones"])),
                "Estado": "Sin coincidencia"
            })
    return pd.DataFrame(resultados)

# Interfaz Streamlit
st.title("Comparador de Facturas y Pedidos (PDF con texto)")

archivo = st.file_uploader("Sube un PDF (con texto reconocible, no imagen)", type=["pdf"])
if archivo:
    texto_total = ""
    with fitz.open(stream=archivo.read(), filetype="pdf") as doc:
        for page in doc:
            texto_total += page.get_text() + "\n"

    productos = extraer_lineas_productos(texto_total)
    df = pd.DataFrame(productos)

    if not df.empty:
        st.subheader("Líneas detectadas")
        st.dataframe(df)

        st.subheader("Resultado de Comparación")
        factura_df = df.iloc[:len(df)//2].copy()
        pedido_df = df.iloc[len(df)//2:].copy()

        factura_grouped = agrupar_por_dimension(factura_df)
        pedido_grouped = agrupar_por_dimension(pedido_df)

        resultado = comparar_factura_pedido(factura_grouped, pedido_grouped)
        st.dataframe(resultado)

        # Exportar a Excel
        output = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativa"
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r in dataframe_to_rows(resultado, index=False, header=True):
            ws.append(r)

        estado_color = {
            "Correcto": "C6EFCE",
            "Precio diferente": "FFFACD",
            "Cantidad diferente": "FFFACD",
            "Ambos diferentes": "FFC7CE",
            "Sin coincidencia": "FFC7CE"
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            estado = row[-1].value
            fill_color = estado_color.get(estado, None)
            if fill_color:
                for cell in row:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        wb.save(output)
        st.download_button("Descargar Excel", data=output.getvalue(), file_name="Comparativa.xlsx")
    else:
        st.warning("No se detectaron líneas válidas. Verifica que el PDF tenga texto seleccionable.")
